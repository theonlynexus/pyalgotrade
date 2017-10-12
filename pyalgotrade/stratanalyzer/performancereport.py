# PyAlgoTrade
#
# Copyright 2011-2015 Gabriel Martin Becedillas Ruiz
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
.. moduleauthor:: Massimo Fierro <massimo.fierro@gmail.com>
"""

from openpyxl import Workbook
from openpyxl.worksheet import Worksheet
from openpyxl.comments.comments import Comment
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, GradientFill, Border, Side
from openpyxl.styles import Alignment, Protection, Font
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.axis import DateAxis

from pyalgotrade.stratanalyzer.extendedtrades import ExtendedTradesAnalyzer


class PerformanceReport(object):
    """Class for computing and generating a TradeStation inspired report.

    This class is dependant on openpyxl to output the report and on
    the ExtendedTradesAnalyzer to have the necessary data.

    :param filename: The filename to save the report as.
    :type trades: :class:`pyalgotrade.stratanalyzer.extendedtrades.ExtendedTradesAnalyzer`.
    """

    def __init__(self):
        pass

    def writeReport(self, filename, trades):
        if not isinstance(trades, ExtendedTradesAnalyzer):
            raise Exception(
                "trades should be an instance of ExtendedTradesAnalyzer")

        wb = Workbook()

        names = wb.sheetnames
        for name in names:
            wb.remove_sheet(wb.get_sheet_by_name(name))

        summarySheet = wb.create_sheet(title="Summary")
        trades_sheet = wb.create_sheet(title="Trades")
        equityGraph_sheet = wb.create_sheet(title="EquityChart")
        detailedEquityGraph_sheet = wb.create_sheet(title="DetailedEquityChart")

        # ----- Trades sheeet -----
        numFormat = "[BLACK][>=0]#,##0.0000;[RED][<0]\\(#,##0.0000\\);General"
        perFormat = "[BLACK][>=0]#0.00%;[RED][<0]\\(#0.00%\);General"

        headerFont = Font(name="Arial", bold=True)
        headerAlign = Alignment(horizontal='center')
        header_fill = PatternFill(
            start_color='AAAAAA', end_color='AAAAAA', fill_type='solid')

        highlightFill = PatternFill(
            start_color='EEEE99', end_color='EEEE99', fill_type='solid')
        highlightBorder = Border(bottom=Side(
            border_style="thin", color="000000"))

        standardFont = Font(name="Arial", size="10")

        for col in range(1, 10):
            trades_sheet.cell(row=1, column=col).font = headerFont
            trades_sheet.cell(row=1, column=col).fill = header_fill
            trades_sheet.cell(row=1, column=col).alignment = headerAlign

        trades_sheet['A1'] = "Trade #\nType"
        trades_sheet['B1'] = "Date"
        trades_sheet['C1'] = "Time"
        trades_sheet['D1'] = "Price"
        trades_sheet['E1'] = "Contracts\nProfit"
        trades_sheet['F1'] = "% Profit\nCum Profit"
        trades_sheet['G1'] = "Run-up\nDrawdown"
        trades_sheet['H1'] = "Entry Eff.\nExit Eff."
        trades_sheet['I1'] = "Total\nEfficiency"

        allTrades = trades.getAll()
        allReturns = trades.getAllReturns()
        allEntryDates = trades.allEnterDates
        allExitDates = trades.allExitDates
        longFlags = trades.allLongFlags
        entryPrices = trades.allEntryPrices
        exitPrices = trades.allExitPrices
        allContracts = trades.allContracts
        allCommissions = trades.getCommissionsForAllTrades()

        excelRow = 2
        cumulativeProfit = 0
        cumulativePnL = 0
        cumulativeLosses = 0

        # ----- Equity graph sheet  -----
        equityGraph_sheet.cell(
            row=1, column=1, value="Trade #")
        equityGraph_sheet.cell(
            row=1, column=2, value="Equity")

        # -----Detailed Equity graph sheet  -----
        detailedEquityGraph_sheet.cell(
            row=1, column=1, value="Timestamp")
        detailedEquityGraph_sheet.cell(
            row=1, column=2, value="Equity")
        r = 2
        for x in sorted(trades.cumPnlDict):
            detailedEquityGraph_sheet.cell(row=r, column=1, value=x)
            detailedEquityGraph_sheet.cell(row=r, column=2, value=trades.cumPnlDict[x])
            r += 1
        # Add chart
        c1 = LineChart()
        c1.title = "Equity curve"
        c1.style = 13
        c1.y_axis.title = 'Equity'
        c1.x_axis.title = 'Date'
        c1.y_axis.auto = True
        c1.y_axis.delete = False
        c1.x_axis = DateAxis(crossAx=100)
        c1.x_axis.number_format = 'd-mmm'
        c1.x_axis.majorTimeUnit = "days"
        c1.x_axis.delete = False
        # c1.x_axis.auto = True        
        c1.legend = None
        x = Reference(
            detailedEquityGraph_sheet,
            min_col=1, min_row=2,
            max_row=len(trades.cumPnlDict) + 2)
        y = Reference(
            detailedEquityGraph_sheet,
            min_col=2, min_row=2,
            max_row=len(trades.cumPnlDict) + 2)
        c1.add_data(y)
        c1.series[0].smooth = False
        c1.series[0].graphicalProperties.line.solidFill = "000000"
        c1.series[0].graphicalProperties.line.width = 10000  # width in EMUs
        c1.width = 30
        c1.height = 15

        detailedEquityGraph_sheet.add_chart(c1, "D3")

        for i in range(0, trades.getCount()):
            # --- Trades sheet ---
            for col in range(1, 10):
                trades_sheet.cell(row=excelRow, column=col).font = standardFont
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).font = standardFont

            trades_sheet.cell(row=excelRow, column=1, value=i + 1)
            trades_sheet.cell(row=excelRow, column=1).alignment = Alignment(
                horizontal='center')
            if longFlags[i]:
                buySell = "Buy"
            else:
                buySell = "Sell"
            trades_sheet.cell(row=excelRow + 1, column=1, value=buySell)
            trades_sheet.cell(
                row=excelRow + 1, column=1).alignment = Alignment(
                    horizontal='center')

            entryDate = allEntryDates[i]
            exitDate = allExitDates[i]
            trades_sheet.cell(row=excelRow, column=2,
                              value=entryDate.strftime("%Y-%m-%d"))
            trades_sheet.cell(row=excelRow + 1, column=2,
                              value=exitDate.strftime("%Y-%m-%d"))

            trades_sheet.cell(row=excelRow, column=3,
                              value=entryDate.strftime("%H:%M"))
            trades_sheet.cell(row=excelRow + 1, column=3,
                              value=exitDate.strftime("%H:%M"))

            trades_sheet.cell(row=excelRow, column=4, value=entryPrices[i])
            trades_sheet.cell(row=excelRow + 1, column=4, value=exitPrices[i])

            trades_sheet.cell(row=excelRow, column=5,
                              value=abs(allContracts[i]))
            trades_sheet.cell(row=excelRow + 1, column=5, value=allTrades[i])
            # TODO(max): Should this include or exclude commissions?
            trades_sheet.cell(
                row=excelRow + 1, column=5).number_format = numFormat

            # TODO(max): Check formula with commissions!
            profitPerc = (
                (exitPrices[i] - allCommissions[i] / allContracts[i]) /
                entryPrices[i]) - 1
            if not longFlags[i]:
                # TODO(max): Check formula with commissions!
                profitPerc = - profitPerc

            trades_sheet.cell(row=excelRow, column=6, value=profitPerc)
            trades_sheet.cell(
                row=excelRow, column=6).number_format = perFormat

            # if longFlags[i]:
            #     profit = (exitPrices[i]-entryPrices[i])*allContracts[i]
            # else:
            #     profit = -(exitPrices[i]-entryPrices[i])*allContracts[i]
            cumulativePnL = cumulativePnL + allTrades[i]
            if allTrades[i] > 0:
                cumulativeProfit = cumulativeProfit + allTrades[i]
            else:
                cumulativeLosses = cumulativeLosses + allTrades[i]
            trades_sheet.cell(row=excelRow + 1, column=6, value=cumulativePnL)
            trades_sheet.cell(
                row=excelRow + 1, column=6).number_format = numFormat

            # Runup & Drawdown
            trades_sheet.cell(row=excelRow, column=7,
                              value=trades.allRunups[i])
            trades_sheet.cell(row=excelRow,
                              column=7).number_format = numFormat
            trades_sheet.cell(row=excelRow + 1, column=7,
                              value=trades.allDrawDowns[i])
            trades_sheet.cell(row=excelRow + 1,
                              column=7).number_format = numFormat

            # Entry & Exit efficiencies
            trades_sheet.cell(row=excelRow, column=8,
                              value=trades.allEntryEff[i])
            trades_sheet.cell(row=excelRow,
                              column=8).number_format = perFormat
            trades_sheet.cell(row=excelRow + 1, column=8,
                              value=trades.allExitEff[i])
            trades_sheet.cell(row=excelRow + 1,
                              column=8).number_format = perFormat

            # Total efficiency
            trades_sheet.cell(row=excelRow + 1, column=9,
                              value=trades.allTotalEff[i])
            trades_sheet.cell(row=excelRow + 1,
                              column=9).number_format = perFormat

            # Set standard font, and highlight style for 2nd row of trade
            for col in range(1, 10):
                # 1st row
                trades_sheet.cell(row=excelRow, column=col).font = standardFont

                # 2nd row
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).font = standardFont
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).fill = highlightFill
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).border = highlightBorder

            excelRow = excelRow + 2

            # ----- Equity graph sheet  -----
            equityGraph_sheet.cell(
                row=i + 2, column=1, value=i + 1)
            equityGraph_sheet.cell(
                row=i + 2, column=2,
                value=trades.initialEquity + cumulativePnL)

        if trades.openPosition is not None:
            pos = trades.openPosition.getPosition()
            if pos != 0:
                i += 1
                for col in range(1, 10):
                    trades_sheet.cell(row=excelRow,
                                      column=col).font = standardFont
                    trades_sheet.cell(row=excelRow + 1,
                                      column=col).font = standardFont
                    trades_sheet.cell(row=excelRow + 1,
                                      column=col).fill = highlightFill
                    trades_sheet.cell(row=excelRow + 1,
                                      column=col).border = highlightBorder

                isLong = trades.openPosition.isLong

                trades_sheet.cell(row=excelRow, column=1, value=i + 1)
                trades_sheet.cell(row=excelRow, column=1).alignment = Alignment(
                    horizontal='center')
                if isLong:
                    buySell = "Buy"
                else:
                    buySell = "Sell"
                trades_sheet.cell(row=excelRow + 1, column=1, value=buySell)
                trades_sheet.cell(
                    row=excelRow + 1, column=1).alignment = Alignment(
                        horizontal='center')

                entryDate = trades.openPosition.entryDate
                trades_sheet.cell(row=excelRow, column=2,
                                  value=entryDate.strftime("%Y-%m-%d"))
                trades_sheet.cell(row=excelRow + 1, column=2,
                                  value=entryDate.strftime("Open"))

                trades_sheet.cell(row=excelRow, column=3,
                                  value=entryDate.strftime("%H:%M"))

                entryPrice = trades.openPosition.entryPrice
                trades_sheet.cell(row=excelRow, column=4, value=entryPrice)
                trades_sheet.cell(row=excelRow + 1, column=4, value="--")

                trades_sheet.cell(row=excelRow, column=5, value=abs(pos))
                trades_sheet.cell(row=excelRow + 1, column=5, value="--")

                trades_sheet.cell(row=excelRow, column=6, value="--")
                trades_sheet.cell(row=excelRow + 1, column=6, value="--")

                trades_sheet.cell(row=excelRow, column=8, value="--")
                trades_sheet.cell(row=excelRow + 1, column=8, value="--")

                trades_sheet.cell(row=excelRow + 1, column=9, value="--")

        # ----- Equity graph sheet -----
        # Add chart
        c1 = LineChart()
        c1.title = "Equity curve"
        c1.style = 13
        c1.y_axis.title = 'Equity'
        c1.x_axis.title = 'Trade #'
        c1.x_axis.scaling.min = 0
        c1.x_axis.scaling.max = len(allTrades) + 3
        c1.x_axis.auto = True
        c1.y_axis.auto = True
        c1.x_axis.delete = False
        c1.y_axis.delete = False
        c1.legend = None
        x = Reference(
            equityGraph_sheet,
            min_col=1, min_row=2,
            max_row=len(allTrades) + 2)
        y = Reference(
            equityGraph_sheet,
            min_col=2, min_row=2,
            max_row=len(allTrades) + 2)
        c1.add_data(y)
        c1.series[0].smooth = False
        c1.series[0].graphicalProperties.line.solidFill = "000000"
        c1.series[0].graphicalProperties.line.width = 10000  # width in EMUs
        c1.width = 30
        c1.height = 15

        equityGraph_sheet.add_chart(c1, "D3")

        # ----- Summary sheeet -----
        titleFont = Font(name="Arial",  size=18, bold=True)
        titleAlign = Alignment(horizontal='center')

        headerFont = Font(name="Arial",  size=14, bold=True)
        headerAlign = Alignment(horizontal='left')

        standardFont = Font(name="Arial", size=10)

        summarySheet['A1'] = "Strategy Performance Report"
        summarySheet.merge_cells("A1:I1")
        summarySheet['A1'].font = titleFont
        summarySheet['A1'].alignment = titleAlign

        summarySheet["B6"] = "Performance Summary: All Trades"
        summarySheet["B6"].font = headerFont
        summarySheet["B6"].alignment = headerAlign

        summarySheet["B8"] = "Net Profits"
        summarySheet["D8"] = cumulativeProfit + cumulativeLosses
        summarySheet["D8"].number_format = numFormat

        summarySheet["F8"] = "Open position P/L"
        summarySheet["H8"] = ""

        summarySheet["B9"] = "Gross Profits"
        summarySheet["D9"] = cumulativeProfit
        summarySheet["D9"].number_format = numFormat
        summarySheet["D9"].comment = Comment(
            "Net profits - Gross losses, i.e. Net profits + Abs(Gross losses)",
            "Report")

        summarySheet["F9"] = "Gross Losses"
        summarySheet["H9"] = cumulativeLosses
        summarySheet["H9"].number_format = numFormat

        summarySheet["B11"] = "Total num. of trades"
        summarySheet["D11"] = trades.getCount()

        summarySheet["F11"] = "Percent profitable"
        if trades.getCount() > 0:
            summarySheet["H11"] = float(
                trades.getProfitableCount()) / float(
                    trades.getCount())
        else:
            summarySheet["H11"] = 0
        summarySheet["H11"].number_format = perFormat

        summarySheet["B12"] = "Num. of winning trades"
        summarySheet["D12"] = trades.getProfitableCount()

        summarySheet["F12"] = "Num. of losing trades"
        summarySheet["H12"] = trades.getUnprofitableCount()

        summarySheet["B14"] = "Largest winning trade"
        if trades.getProfitableCount() > 0:
            summarySheet["D14"] = allTrades.max()
        else:
            summarySheet["D14"] = 0
        summarySheet["D14"].number_format = numFormat

        summarySheet["F14"] = "Largest losing trade"
        if trades.getUnprofitableCount() > 0:
            summarySheet["H14"] = allTrades.min()
        else:
            summarySheet["H14"] = 0
        summarySheet["H14"].number_format = numFormat

        def negativeToZero(x):
            if x > 0:
                return x
            else:
                return 0

        def positiveToZero(x):
            if x >= 0:
                return 0
            else:
                return x

        avgWin = 0
        avgLoss = 0
        if trades.getProfitableCount() > 0:
            avgWin = cumulativeProfit / trades.getProfitableCount()
        summarySheet["B15"] = "Average winning trade"
        summarySheet["D15"] = avgWin
        summarySheet["D15"].number_format = numFormat

        if trades.getUnprofitableCount() > 0:
            avgLoss = cumulativeLosses / trades.getUnprofitableCount()
        summarySheet["F15"] = "Average losing trade"
        summarySheet["H15"] = avgLoss
        summarySheet["H15"].number_format = numFormat

        summarySheet["B16"] = "Ratio avg. win/avg. loss"
        if avgLoss != 0:
            summarySheet["D16"] = - avgWin / avgLoss
        else:
            summarySheet["D16"] = 'NaN'
        summarySheet["D16"].number_format = numFormat

        summarySheet["F16"] = "Avg trade (win & loss)"
        if trades.getCount() > 0:
            summarySheet["H16"] = cumulativePnL / trades.getCount()
        else:
            summarySheet["H16"] = 0
        summarySheet["H16"].number_format = numFormat

        summarySheet["B21"] = "Max intraday drawdown"
        summarySheet["D21"] = ""

        summarySheet["B22"] = "Profit factor"
        if cumulativeLosses != 0:
            summarySheet["D22"] = - cumulativeProfit / cumulativeLosses
            summarySheet["D22"].number_format = numFormat
        else:
            summarySheet["D22"] = "Inf"
        summarySheet["D22"].comment = Comment(
            "- Gross profits / Gross losses", "Report")

        summarySheet["F22"] = "Max contracts held"
        summarySheet["H22"] = ""

        summarySheet["B23"] = "Account size required"
        summarySheet["D23"] = ""  # ABS(max intraday drawdown)

        summarySheet["F23"] = "Return on account"
        summarySheet["H23"] = ""  # net profit / account size required

        # Save the file
        wb.save(filename)
